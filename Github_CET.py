"""
Build a GitHub + GitHub Copilot Cost Estimate Tool (CET) workbook.

Usage:
    pip install openpyxl
    python build_cet.py [output.xlsx]

Produces a 7-sheet workbook: Read_Me, Dashboard, Licences, Platform_Usage,
Summary, SKU_Rate_Card, Plan_Allowances.

All rates are GitHub public list prices in USD, verified 27 August 2026.
Edit the RATECARD / PLANS / METERED / LICENCES tables below to change them.

NOTE: openpyxl writes formulas without cached values, so every formula cell
reads back as None until the file is opened in Excel/LibreOffice once (or run
through a headless recalc). Excel calculates everything on open.
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

FONT = "Arial"

# ---------------------------------------------------------------- styles
def F(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

TITLE = F(16, True, "24292F")
SUBTITLE = F(10, False, "57606A", italic=True)
SECTION = F(11, True, "FFFFFF")
HDR = F(10, True, "FFFFFF")
BODY = F(10)
LINK = F(10, False, "008000")          # green = link to another sheet
INPUT_F = F(10, False, "0000FF")       # blue = hardcoded input
BOLD = F(10, True)
NOTE = F(9, False, "57606A", italic=True)

SECTION_FILL = PatternFill("solid", fgColor="24292F")
HDR_FILL = PatternFill("solid", fgColor="4A5568")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
GROUP_FILL = PatternFill("solid", fgColor="E7EDF3")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
BAND_FILL = PatternFill("solid", fgColor="F6F8FA")

thin = Side(style="thin", color="BFC7CF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0 = '$#,##0;($#,##0);"-"'
RATE = '$#,##0.0000;($#,##0.0000);"-"'
QTY = '#,##0;(#,##0);"-"'
PCT = '0.0%'

def put(ws, cell, value, font=BODY, fmt=None, fill=None, align=None, border=False, wrap=False):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c

def band(ws, row, first_col, last_col, fill, font=None):
    for i in range(first_col, last_col + 1):
        c = ws.cell(row=row, column=i)
        c.fill = fill
        if font:
            c.font = font

# ---------------------------------------------------------------- data
# (category, product, sku, description, unit_type, list_rate, note)
RATECARD = [
    ("Platform licences", "github_platform", "github_free", "GitHub Free (individual)", "user-months", 0.00, "github.com/pricing"),
    ("Platform licences", "github_platform", "github_pro", "GitHub Pro (individual)", "user-months", 4.00, "github.com/pricing"),
    ("Platform licences", "github_platform", "github_free_org", "GitHub Free for organizations", "user-months", 0.00, "github.com/pricing"),
    ("Platform licences", "github_platform", "github_team", "GitHub Team", "user-months", 4.00, "github.com/pricing"),
    ("Platform licences", "github_platform", "ghec_licenses", "GitHub Enterprise Cloud licence", "user-months", 21.00, "github.com/pricing; SKU per GitHub billing reference"),

    ("Copilot licences", "copilot", "copilot_for_business", "Copilot Business (1,900 AI credits/user/mo)", "user-months", 19.00, "docs.github.com/copilot/get-started/plans"),
    ("Copilot licences", "copilot", "copilot_enterprise", "Copilot Enterprise (3,900 AI credits/user/mo)", "user-months", 39.00, "docs.github.com/copilot/get-started/plans"),
    ("Copilot licences", "copilot", "copilot_pro", "Copilot Pro (1,500 AI credits/mo)", "user-months", 10.00, "Billed under SKU copilot_standalone"),
    ("Copilot licences", "copilot", "copilot_pro_plus", "Copilot Pro+ (7,000 AI credits/mo)", "user-months", 39.00, "Billed under SKU copilot_standalone"),
    ("Copilot licences", "copilot", "copilot_max", "Copilot Max (20,000 AI credits/mo)", "user-months", 100.00, "Billed under SKU copilot_standalone"),

    ("Security & quality licences", "ghas", "ghas_secret_protection_licenses", "GitHub Secret Protection", "committer-months", 19.00, "Per active committer, 90-day window"),
    ("Security & quality licences", "ghas", "ghas_code_security_licenses", "GitHub Code Security", "committer-months", 30.00, "Per active committer, 90-day window"),
    ("Security & quality licences", "ghas", "ghas_licenses", "GitHub Advanced Security (both products)", "committer-months", 49.00, "Use instead of the two rows above, not as well as"),
    ("Security & quality licences", "code_quality", "code_quality_licenses", "GitHub Code Quality (GA 20 Jul 2026)", "committer-months", 10.00, "Per active committer on enabled repos"),

    ("Actions - standard runners", "actions", "actions_linux_slim", "Linux 1-core x64 (ubuntu-slim)", "minutes", 0.002, "Jan-2026 rates; billed per whole minute"),
    ("Actions - standard runners", "actions", "actions_linux", "Linux 2-core x64 (ubuntu-latest)", "minutes", 0.006, "Jan-2026 rates; billed per whole minute"),
    ("Actions - standard runners", "actions", "actions_linux_arm", "Linux 2-core arm64", "minutes", 0.005, "Jan-2026 rates"),
    ("Actions - standard runners", "actions", "actions_windows", "Windows 2-core x64", "minutes", 0.010, "Consumes included minutes at 2x"),
    ("Actions - standard runners", "actions", "actions_windows_arm", "Windows 2-core arm64", "minutes", 0.010, "Consumes included minutes at 2x"),
    ("Actions - standard runners", "actions", "actions_macos", "macOS 3 or 4-core (M1 / Intel)", "minutes", 0.062, "Consumes included minutes at 10x"),
    ("Actions - standard runners", "actions", "actions_self_hosted_linux", "Self-hosted Linux runner", "minutes", 0.000, "No GitHub per-minute charge; Dec-2025 $0.002 platform fee was postponed"),
    ("Actions - standard runners", "actions", "actions_self_hosted_windows", "Self-hosted Windows runner", "minutes", 0.000, "No GitHub per-minute charge; you still pay for your own compute"),

    ("Actions - larger runners (x64)", "actions", "actions_linux_2_core_advanced", "Linux Advanced 2-core", "minutes", 0.006, "Larger runners cannot use included minutes"),
    ("Actions - larger runners (x64)", "actions", "actions_linux_4_core", "Linux 4-core", "minutes", 0.012, "Larger runners cannot use included minutes"),
    ("Actions - larger runners (x64)", "actions", "actions_linux_8_core", "Linux 8-core", "minutes", 0.022, ""),
    ("Actions - larger runners (x64)", "actions", "actions_linux_16_core", "Linux 16-core", "minutes", 0.042, ""),
    ("Actions - larger runners (x64)", "actions", "actions_linux_32_core", "Linux 32-core", "minutes", 0.082, ""),
    ("Actions - larger runners (x64)", "actions", "actions_linux_64_core", "Linux 64-core", "minutes", 0.162, ""),
    ("Actions - larger runners (x64)", "actions", "actions_linux_96_core", "Linux 96-core", "minutes", 0.252, ""),
    ("Actions - larger runners (x64)", "actions", "actions_windows_4_core", "Windows 4-core", "minutes", 0.022, ""),
    ("Actions - larger runners (x64)", "actions", "actions_windows_8_core", "Windows 8-core", "minutes", 0.042, ""),
    ("Actions - larger runners (x64)", "actions", "actions_windows_16_core", "Windows 16-core", "minutes", 0.082, ""),
    ("Actions - larger runners (x64)", "actions", "actions_windows_32_core", "Windows 32-core", "minutes", 0.162, ""),
    ("Actions - larger runners (x64)", "actions", "actions_windows_64_core", "Windows 64-core", "minutes", 0.322, ""),
    ("Actions - larger runners (x64)", "actions", "actions_windows_96_core", "Windows 96-core", "minutes", 0.552, ""),
    ("Actions - larger runners (x64)", "actions", "actions_macos_l", "macOS 12-core (macos_l)", "minutes", 0.077, ""),

    ("Actions - larger runners (arm64)", "actions", "actions_linux_2_core_arm", "Linux 2-core arm64 (larger)", "minutes", 0.005, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_linux_4_core_arm", "Linux 4-core arm64", "minutes", 0.008, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_linux_8_core_arm", "Linux 8-core arm64", "minutes", 0.014, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_linux_16_core_arm", "Linux 16-core arm64", "minutes", 0.026, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_linux_32_core_arm", "Linux 32-core arm64", "minutes", 0.050, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_linux_64_core_arm", "Linux 64-core arm64", "minutes", 0.098, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_windows_2_core_arm", "Windows 2-core arm64 (larger)", "minutes", 0.008, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_windows_4_core_arm", "Windows 4-core arm64", "minutes", 0.014, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_windows_8_core_arm", "Windows 8-core arm64", "minutes", 0.026, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_windows_16_core_arm", "Windows 16-core arm64", "minutes", 0.050, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_windows_32_core_arm", "Windows 32-core arm64", "minutes", 0.098, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_windows_64_core_arm", "Windows 64-core arm64", "minutes", 0.194, ""),
    ("Actions - larger runners (arm64)", "actions", "actions_macos_xl", "macOS 5-core M2 Pro (macos_xl)", "minutes", 0.102, ""),

    ("Actions - GPU runners", "actions", "actions_linux_4_core_gpu", "Linux 4-core GPU", "minutes", 0.052, ""),
    ("Actions - GPU runners", "actions", "actions_windows_4_core_gpu", "Windows 4-core GPU", "minutes", 0.102, ""),

    ("Actions - storage", "actions", "actions_storage", "Actions artifact & log storage", "gigabyte-months", 0.25, "Metered in GB-hours (~$0.000342/GB-hour); allowance shared with Packages"),
    ("Actions - storage", "actions", "actions_cache_storage", "Actions cache storage", "gigabyte-months", 0.07, "10 GB included per repository on GHEC"),
    ("Actions - storage", "actions", "actions_custom_image_storage", "Actions custom runner image storage", "gigabyte-months", 0.07, "150 GB included on GHEC"),

    ("Packages", "packages", "packages_storage", "Packages storage", "gigabyte-months", 0.25, "Allowance shared with Actions storage"),
    ("Packages", "packages", "packages_bandwidth", "Packages data transfer out", "gigabytes", 0.50, "Free for public repos and inside Actions workflows"),

    ("Git LFS", "git_lfs", "git_lfs_storage", "Git LFS storage", "gigabyte-months", 0.07, "Metered hourly, billed per GiB-month"),
    ("Git LFS", "git_lfs", "git_lfs_bandwidth", "Git LFS download bandwidth", "gigabytes", 0.0875, "Every clone / fetch / CI pull counts"),

    ("Codespaces", "codespaces", "codespaces_compute_d2", "Codespaces compute 2-core", "hours", 0.18, "$0.09 per core-hour; 2-core burns 2 core-hours per hour"),
    ("Codespaces", "codespaces", "codespaces_compute_d4", "Codespaces compute 4-core", "hours", 0.36, ""),
    ("Codespaces", "codespaces", "codespaces_compute_d8", "Codespaces compute 8-core", "hours", 0.72, ""),
    ("Codespaces", "codespaces", "codespaces_compute_d16", "Codespaces compute 16-core", "hours", 1.44, ""),
    ("Codespaces", "codespaces", "codespaces_compute_d32", "Codespaces compute 32-core", "hours", 2.88, ""),
    ("Codespaces", "codespaces", "codespaces_storage", "Codespaces storage", "gigabyte-months", 0.07, "Charged while stopped, until deleted"),
    ("Codespaces", "codespaces", "codespaces_prebuild_storage", "Codespaces prebuild storage", "gigabyte-months", 0.07, ""),

    ("AI credits", "copilot", "copilot_ai_credits", "Copilot AI credits (chat, agents, code review)", "ai-credits", 0.01, "1 AI credit = $0.01. Completions & next-edit suggestions are not billed"),
    ("AI credits", "code_quality", "code_quality_ai_credit", "Code Quality AI credits (autofix, AI findings)", "ai-credits", 0.01, "Drawn from the shared AI credit pool; no model switching"),
    ("AI credits", "spark", "spark_ai_credits", "GitHub Spark AI credits", "ai-credits", 0.01, "Dedicated Spark SKU since Nov 2025"),
    ("AI credits", "models", "models_inference", "GitHub Models inference (blended)", "million-tokens", 1.00, "MODEL-SPECIFIC - replace with your own blended rate"),

    ("Copilot cloud sandboxes", "sandbox", "sandbox_linux", "Cloud sandbox compute", "compute-hours", 0.0864, "GitHub publishes $0.000024 per compute-second"),
    ("Copilot cloud sandboxes", "sandbox", "sandbox_memory", "Cloud sandbox memory", "GiB-hours", 0.0108, "GitHub publishes $0.000003 per GiB-second"),
    ("Copilot cloud sandboxes", "sandbox", "sandbox_snapshot", "Cloud sandbox snapshot storage", "GiB-months", 0.005, "Charged from stop until delete"),
]

# metered SKUs shown on Platform_Usage: (sku, allowance_formula_or_None, example_qty)
METERED = [
    ("actions_linux_slim", None, 0),
    ("actions_linux", "PLAN_MIN", 120000),
    ("actions_linux_arm", None, 0),
    ("actions_windows", None, 15000),
    ("actions_windows_arm", None, 0),
    ("actions_macos", None, 2000),
    ("actions_self_hosted_linux", None, 0),
    ("actions_self_hosted_windows", None, 0),
    ("actions_linux_2_core_advanced", None, 0),
    ("actions_linux_4_core", None, 8000),
    ("actions_linux_8_core", None, 0),
    ("actions_linux_16_core", None, 0),
    ("actions_linux_32_core", None, 0),
    ("actions_linux_64_core", None, 0),
    ("actions_linux_96_core", None, 0),
    ("actions_windows_4_core", None, 0),
    ("actions_windows_8_core", None, 0),
    ("actions_windows_16_core", None, 0),
    ("actions_windows_32_core", None, 0),
    ("actions_windows_64_core", None, 0),
    ("actions_windows_96_core", None, 0),
    ("actions_macos_l", None, 0),
    ("actions_linux_2_core_arm", None, 0),
    ("actions_linux_4_core_arm", None, 0),
    ("actions_linux_8_core_arm", None, 0),
    ("actions_linux_16_core_arm", None, 0),
    ("actions_linux_32_core_arm", None, 0),
    ("actions_linux_64_core_arm", None, 0),
    ("actions_windows_2_core_arm", None, 0),
    ("actions_windows_4_core_arm", None, 0),
    ("actions_windows_8_core_arm", None, 0),
    ("actions_windows_16_core_arm", None, 0),
    ("actions_windows_32_core_arm", None, 0),
    ("actions_windows_64_core_arm", None, 0),
    ("actions_macos_xl", None, 0),
    ("actions_linux_4_core_gpu", None, 0),
    ("actions_windows_4_core_gpu", None, 0),
    ("actions_storage", "PLAN_ASTOR", 120),
    ("actions_cache_storage", None, 250),
    ("actions_custom_image_storage", None, 0),
    ("packages_storage", "PLAN_PSTOR", 200),
    ("packages_bandwidth", "PLAN_PBAND", 300),
    ("git_lfs_storage", "PLAN_LFSSTOR", 400),
    ("git_lfs_bandwidth", "PLAN_LFSBAND", 600),
    ("codespaces_compute_d2", "PLAN_CSHRS", 0),
    ("codespaces_compute_d4", None, 500),
    ("codespaces_compute_d8", None, 0),
    ("codespaces_compute_d16", None, 0),
    ("codespaces_compute_d32", None, 0),
    ("codespaces_storage", "PLAN_CSSTOR", 600),
    ("codespaces_prebuild_storage", None, 0),
    ("copilot_ai_credits", "AI_POOL", 450000),
    ("code_quality_ai_credit", None, 0),
    ("spark_ai_credits", None, 0),
    ("models_inference", None, 0),
    ("sandbox_linux", None, 0),
    ("sandbox_memory", None, 0),
    ("sandbox_snapshot", None, 0),
]

# Licence rows: (product, sku_or_formula, label, unit, example_qty, credits_per_seat)
LICENCES = [
    ("github_platform", "PLAN", "GitHub platform licences (plan set on Dashboard)", "user-months", 250, None),
    ("copilot", "copilot_for_business", "Copilot Business", "user-months", 200, 1900),
    ("copilot", "copilot_enterprise", "Copilot Enterprise", "user-months", 0, 3900),
    ("copilot", "copilot_pro", "Copilot Pro (individual)", "user-months", 0, 1500),
    ("copilot", "copilot_pro_plus", "Copilot Pro+ (individual)", "user-months", 0, 7000),
    ("copilot", "copilot_max", "Copilot Max (individual)", "user-months", 0, 20000),
    ("ghas", "ghas_secret_protection_licenses", "GitHub Secret Protection", "committer-months", 150, None),
    ("ghas", "ghas_code_security_licenses", "GitHub Code Security", "committer-months", 150, None),
    ("ghas", "ghas_licenses", "GitHub Advanced Security bundle (alternative to the two rows above)", "committer-months", 0, None),
    ("code_quality", "code_quality_licenses", "GitHub Code Quality", "committer-months", 0, None),
]

PLANS = [
    # plan, sku, minutes, actions_stor, pkg_stor, pkg_band, lfs_stor, lfs_band, cs_core_hrs, cs_stor
    ("GitHub Free", "github_free", 2000, 0.5, 0.5, 1, 10, 10, 120, 15),
    ("GitHub Pro", "github_pro", 3000, 1, 2, 10, 10, 10, 180, 20),
    ("GitHub Free for organizations", "github_free_org", 2000, 0.5, 0.5, 1, 10, 10, 0, 0),
    ("GitHub Team", "github_team", 3000, 2, 2, 10, 250, 250, 0, 0),
    ("GitHub Enterprise Cloud", "ghec_licenses", 50000, 50, 50, 100, 250, 250, 0, 0),
]

SUMMARY_PRODUCTS = [
    ("github_platform", "GitHub platform licences"),
    ("copilot", "GitHub Copilot (seats + AI credits)"),
    ("ghas", "GitHub Advanced Security"),
    ("code_quality", "GitHub Code Quality"),
    ("actions", "GitHub Actions (compute + storage)"),
    ("packages", "GitHub Packages"),
    ("git_lfs", "Git LFS"),
    ("codespaces", "GitHub Codespaces"),
    ("sandbox", "Copilot cloud sandboxes"),
    ("models", "GitHub Models"),
    ("spark", "GitHub Spark"),
]

wb = Workbook()

# ================================================================ Read_Me
ws = wb.active
ws.title = "Read_Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 96

put(ws, "B2", "GitHub & GitHub Copilot - Cost Estimate Tool (CET)", TITLE)
put(ws, "B3", "List prices in USD, verified against GitHub public documentation on 27 August 2026.", SUBTITLE)

rows = [
    ("How to use it", "Fill in the yellow cells only. Everything else is calculated. Work top to bottom: Dashboard -> Licences -> Platform_Usage."),
    ("1. Dashboard", "Enter the customer name, term, discount and contingency, and pick the GitHub platform plan. Read the totals and the cost-by-product chart here."),
    ("2. Licences", "Enter seat counts and active-committer counts. Seat rates come from the rate card automatically."),
    ("3. Platform_Usage", "Enter expected monthly usage against each metered SKU. Included plan allowances are deducted for you."),
    ("4. Summary", "Cost by product, per month, per year and across the whole term."),
    ("5. SKU_Rate_Card", "Every GitHub and Copilot SKU with its unit and list rate. Overwrite a list rate here to model negotiated pricing."),
    ("6. Plan_Allowances", "Included usage that comes with each GitHub plan. Drives the 'Included' column on Platform_Usage."),
    ("", ""),
    ("Colour legend", "YELLOW FILL = type here.  Blue text = a number you can change.  Black text = a formula.  Green text = a link to another sheet."),
    ("", ""),
    ("Worked example", "The workbook opens pre-filled with a 250-developer GitHub Enterprise Cloud scenario (200 Copilot Business seats, 150 security committers) so you can see the expected format. Overwrite every yellow cell with your own numbers."),
    ("", ""),
    ("How the maths works", "Billable qty = monthly usage - included allowance (floored at zero).  Line cost = billable qty x applied rate.  Applied rate = list rate x (1 - discount).  Term cost = monthly total x term months."),
    ("Included minutes", "GitHub's included Actions minutes are a single Linux-equivalent pool: Windows burns them at 2x and macOS at 10x. The tool shows the whole pool against the standard Linux runner. If your mix is Windows or macOS heavy, reduce that allowance manually on Plan_Allowances."),
    ("Larger runners", "Larger runners (4-core and above), GPU runners and macOS larger runners can never draw on included minutes. They are billed from the first minute."),
    ("Self-hosted runners", "Shown at $0.000/minute. GitHub announced a $0.002/minute platform charge in December 2025 and postponed it before it took effect; it is not live as of August 2026. Your own infrastructure cost is not included in this tool."),
    ("Storage units", "GitHub meters storage in gigabyte-hours. This tool asks for average GB held per month and applies the published per-GB-month rate, which is the same answer with far less arithmetic."),
    ("AI credits", "1 AI credit = $0.01. Each Copilot seat contributes credits to one shared pool (Business 1,900/user/mo, Enterprise 3,900). Overage is billed per credit. Code completions and next-edit suggestions are unlimited and never billed."),
    ("", ""),
    ("SKU naming", "SKU IDs follow GitHub's current billing reference. Some usage exports still show older singular forms - action_storage, package_storage, copilot_ai_credit, action_self_hosted_linux - which map to the same meters."),
    ("Not included", "Taxes, GitHub Enterprise Server, self-hosted runner infrastructure, GitHub Marketplace apps, Sponsors, professional services, and any Microsoft EA or volume discount you have negotiated."),
    ("Health check", "Rates move. Re-check docs.github.com/billing before you send an estimate to a customer."),
]
r = 5
for label, text in rows:
    if label:
        put(ws, f"B{r}", label, BOLD)
        put(ws, f"C{r}", text, BODY, wrap=True)
        ws.row_dimensions[r].height = 30 if len(text) > 110 else 15
    r += 1

put(ws, f"B{r+1}", "Sources", BOLD)
srcs = [
    "docs.github.com/en/billing/reference/actions-runner-pricing (runner per-minute rates)",
    "docs.github.com/en/billing/reference/product-and-sku-names (SKU identifiers)",
    "docs.github.com/en/billing/reference/product-usage-included (included allowances by plan)",
    "docs.github.com/en/copilot/get-started/plans (Copilot plan prices and AI credit allowances)",
    "docs.github.com/en/billing/concepts/product-billing/cloud-and-local-sandboxes (sandbox meters)",
    "github.com/pricing and github.com/pricing/calculator (plan seats, storage and transfer rates)",
    "github.blog changelog, 16 Jun 2026 (Code Quality GA at $10 per active committer)",
]
r = r + 2
for s in srcs:
    put(ws, f"C{r}", s, NOTE)
    r += 1

# ================================================================ SKU_Rate_Card
rc = wb.create_sheet("SKU_Rate_Card")
rc.sheet_view.showGridLines = False
widths = {"A": 30, "B": 16, "C": 34, "D": 44, "E": 17, "F": 14, "G": 16, "H": 62}
for k, v in widths.items():
    rc.column_dimensions[k].width = v

put(rc, "A1", "SKU Rate Card - GitHub & GitHub Copilot (USD list prices)", TITLE)
put(rc, "A2", "Blue list rates are editable. applied_cost_per_quantity = list rate x (1 - discount on the Dashboard).", SUBTITLE)

RC_HDR = 4
hdrs = ["Category", "Product", "sku", "Description", "unit_type", "list_rate_usd",
        "applied_cost_per_quantity", "Notes / source"]
for i, h in enumerate(hdrs, start=1):
    put(rc, f"{get_column_letter(i)}{RC_HDR}", h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
rc.row_dimensions[RC_HDR].height = 30

RC_FIRST = RC_HDR + 1
sku_row = {}
r = RC_FIRST
prev_cat = None
for cat, prod, sku, desc, unit, rate, note in RATECARD:
    put(rc, f"A{r}", cat, BODY, border=True)
    put(rc, f"B{r}", prod, BODY, border=True)
    put(rc, f"C{r}", sku, BODY, border=True)
    put(rc, f"D{r}", desc, BODY, border=True)
    put(rc, f"E{r}", unit, BODY, border=True, align="center")
    put(rc, f"F{r}", rate, INPUT_F, fmt=RATE, fill=INPUT_FILL, border=True)
    put(rc, f"G{r}", f"=F{r}*(1-Dashboard!$C$11)", BODY, fmt=RATE, border=True)
    put(rc, f"H{r}", note, NOTE, border=True)
    if cat != prev_cat:
        band(rc, r, 1, 1, GROUP_FILL, BOLD)
        prev_cat = cat
    sku_row[sku] = r
    r += 1
RC_LAST = r - 1
rc.freeze_panes = "A5"
rc.auto_filter.ref = f"A{RC_HDR}:H{RC_LAST}"

RATE_LOOKUP = ("IFERROR(INDEX(SKU_Rate_Card!$G${f}:$G${l},"
               "MATCH({key},SKU_Rate_Card!$C${f}:$C${l},0)),0)").format(f=RC_FIRST, l=RC_LAST, key="{key}")

# ================================================================ Plan_Allowances
pa = wb.create_sheet("Plan_Allowances")
pa.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJK", [32, 20, 14, 16, 15, 15, 16, 15, 16, 17, 16]):
    pa.column_dimensions[col].width = w

put(pa, "A1", "Included usage by GitHub plan", TITLE)
put(pa, "A2", "Source: docs.github.com/en/billing/reference/product-usage-included. Edit the blue cells to model a custom entitlement.", SUBTITLE)

PA_HDR = 4
pa_hdrs = ["Plan", "Platform sku", "Seat rate (list)", "Actions minutes / mo",
           "Actions storage GB", "Packages storage GB", "Packages transfer GB / mo",
           "Git LFS storage GB", "Git LFS bandwidth GB / mo", "Codespaces core-hours / mo",
           "Codespaces storage GB"]
for i, h in enumerate(pa_hdrs, start=1):
    put(pa, f"{get_column_letter(i)}{PA_HDR}", h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
pa.row_dimensions[PA_HDR].height = 42

PA_FIRST = PA_HDR + 1
r = PA_FIRST
for plan, sku, mins, astor, pstor, pband, lstor, lband, cshrs, csstor in PLANS:
    put(pa, f"A{r}", plan, BODY, border=True)
    put(pa, f"B{r}", sku, BODY, border=True)
    put(pa, f"C{r}", "=" + RATE_LOOKUP.format(key=f"$B{r}").replace("$G$", "$F$"), BODY, fmt=MONEY, border=True)
    for i, v in zip("DEFGHIJK", [mins, astor, pstor, pband, lstor, lband, cshrs, csstor]):
        put(pa, f"{i}{r}", v, INPUT_F, fmt=QTY, fill=INPUT_FILL, border=True)
    r += 1
PA_LAST = r - 1

PA_SEL = PA_LAST + 3
put(pa, f"A{PA_SEL-1}", "Plan selected on the Dashboard", SECTION, fill=SECTION_FILL)
band(pa, PA_SEL - 1, 1, 11, SECTION_FILL, SECTION)
put(pa, f"A{PA_SEL}", "=Dashboard!$C$9", LINK, border=True)
for i in "BCDEFGHIJK":
    fmt = MONEY if i == "C" else QTY
    put(pa, f"{i}{PA_SEL}",
        f"=IFERROR(INDEX({i}${PA_FIRST}:{i}${PA_LAST},MATCH($A${PA_SEL},$A${PA_FIRST}:$A${PA_LAST},0)),0)",
        BODY, fmt=fmt, border=True)
band(pa, PA_SEL, 1, 11, TOTAL_FILL)
put(pa, f"A{PA_SEL+2}", "Platform_Usage reads its 'Included' column from this row.", NOTE)

# ================================================================ Licences
lic = wb.create_sheet("Licences")
lic.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJ", [17, 33, 52, 18, 14, 14, 15, 15, 17, 52]):
    lic.column_dimensions[col].width = w

put(lic, "A1", "Licences & subscriptions", TITLE)
put(lic, "A2", "Enter seats and active committers in the yellow column. Set the platform plan on the Dashboard.", SUBTITLE)

L_HDR = 4
l_hdrs = ["Product", "sku", "Plan / description", "unit_type", "Qty (seats)",
          "applied_cost_per_quantity", "Monthly cost", "Annual cost",
          "AI credits per seat / mo", "Notes"]
for i, h in enumerate(l_hdrs, start=1):
    put(lic, f"{get_column_letter(i)}{L_HDR}", h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
lic.row_dimensions[L_HDR].height = 32

L_FIRST = L_HDR + 1
r = L_FIRST
for prod, sku, label, unit, qty, credits in LICENCES:
    put(lic, f"A{r}", prod, BODY, border=True)
    if sku == "PLAN":
        put(lic, f"B{r}", f"=Plan_Allowances!$B${PA_SEL}", LINK, border=True)
    else:
        put(lic, f"B{r}", sku, BODY, border=True)
    put(lic, f"C{r}", label, BODY, border=True)
    put(lic, f"D{r}", unit, BODY, border=True, align="center")
    put(lic, f"E{r}", qty, INPUT_F, fmt=QTY, fill=INPUT_FILL, border=True)
    put(lic, f"F{r}", "=" + RATE_LOOKUP.format(key=f"$B{r}"), BODY, fmt=MONEY, border=True)
    put(lic, f"G{r}", f"=E{r}*F{r}", BODY, fmt=MONEY, border=True)
    put(lic, f"H{r}", f"=G{r}*12", BODY, fmt=MONEY, border=True)
    if credits:
        put(lic, f"I{r}", credits, INPUT_F, fmt=QTY, fill=INPUT_FILL, border=True)
    else:
        put(lic, f"I{r}", 0, BODY, fmt=QTY, border=True)
    note = ""
    if sku == "PLAN":
        note = "Plan and rate follow the Dashboard selection"
    elif sku == "ghas_licenses":
        note = "Use this OR the two rows above, never both"
    elif prod == "ghas" or prod == "code_quality":
        note = "Active committers in the last 90 days, not total seats"
    elif credits:
        note = "Seats contribute to the shared AI credit pool below"
    put(lic, f"J{r}", note, NOTE, border=True)
    r += 1
L_LAST = r - 1

L_TOT = L_LAST + 1
put(lic, f"A{L_TOT}", "Total licences", BOLD, border=True)
put(lic, f"E{L_TOT}", f"=SUM(E{L_FIRST}:E{L_LAST})", BOLD, fmt=QTY, border=True)
put(lic, f"G{L_TOT}", f"=SUM(G{L_FIRST}:G{L_LAST})", BOLD, fmt=MONEY, border=True)
put(lic, f"H{L_TOT}", f"=SUM(H{L_FIRST}:H{L_LAST})", BOLD, fmt=MONEY, border=True)
band(lic, L_TOT, 1, 10, TOTAL_FILL)

L_POOL = L_TOT + 2
put(lic, f"C{L_POOL}", "Included Copilot AI credits per month (shared pool)", BOLD)
put(lic, f"G{L_POOL}", f"=SUMPRODUCT(E{L_FIRST}:E{L_LAST},I{L_FIRST}:I{L_LAST})", BOLD, fmt=QTY, border=True)
put(lic, f"H{L_POOL}", "credits / month - deducted on Platform_Usage", NOTE)
lic.freeze_panes = "A5"


# ================================================================ Platform_Usage
pu = wb.create_sheet("Platform_Usage")
pu.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJK", [31, 14, 32, 17, 16, 15, 15, 17, 15, 15, 56]):
    pu.column_dimensions[col].width = w

put(pu, "A1", "Platform usage - metered products", TITLE)
put(pu, "A2", "Enter expected monthly usage in the yellow column. Included allowances come from the plan chosen on the Dashboard.", SUBTITLE)

U_HDR = 4
u_hdrs = ["Category", "Product", "sku", "unit_type", "Qty per month",
          "Included per month", "Billable qty", "applied_cost_per_quantity",
          "Monthly cost", "Annual cost", "Notes"]
for i, h in enumerate(u_hdrs, start=1):
    put(pu, f"{get_column_letter(i)}{U_HDR}", h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
pu.row_dimensions[U_HDR].height = 32

ALLOW = {
    "PLAN_MIN": f"=Plan_Allowances!$D${PA_SEL}",
    "PLAN_ASTOR": f"=Plan_Allowances!$E${PA_SEL}",
    "PLAN_PSTOR": f"=Plan_Allowances!$F${PA_SEL}",
    "PLAN_PBAND": f"=Plan_Allowances!$G${PA_SEL}",
    "PLAN_LFSSTOR": f"=Plan_Allowances!$H${PA_SEL}",
    "PLAN_LFSBAND": f"=Plan_Allowances!$I${PA_SEL}",
    "PLAN_CSHRS": f"=Plan_Allowances!$J${PA_SEL}/2",
    "PLAN_CSSTOR": f"=Plan_Allowances!$K${PA_SEL}",
    "AI_POOL": f"=Licences!$G${L_POOL}",
}

meta = {row[2]: row for row in RATECARD}
U_FIRST = U_HDR + 1
r = U_FIRST
prev_cat = None
for sku, allow, qty in METERED:
    cat, prod, _sku, desc, unit, rate, note = meta[sku]
    put(pu, f"A{r}", cat, BODY, border=True)
    put(pu, f"B{r}", prod, BODY, border=True)
    put(pu, f"C{r}", sku, BODY, border=True)
    put(pu, f"D{r}", unit, BODY, border=True, align="center")
    put(pu, f"E{r}", qty, INPUT_F, fmt=QTY, fill=INPUT_FILL, border=True)
    if allow:
        put(pu, f"F{r}", ALLOW[allow], LINK, fmt=QTY, border=True)
    else:
        put(pu, f"F{r}", 0, BODY, fmt=QTY, border=True)
    put(pu, f"G{r}", f"=MAX(0,E{r}-F{r})", BODY, fmt=QTY, border=True)
    put(pu, f"H{r}", "=" + RATE_LOOKUP.format(key=f"$C{r}"), BODY, fmt=RATE, border=True)
    put(pu, f"I{r}", f"=G{r}*H{r}", BODY, fmt=MONEY, border=True)
    put(pu, f"J{r}", f"=I{r}*12", BODY, fmt=MONEY, border=True)
    put(pu, f"K{r}", desc + (" - " + note if note else ""), NOTE, border=True)
    if cat != prev_cat:
        band(pu, r, 1, 1, GROUP_FILL, BOLD)
        prev_cat = cat
    r += 1
U_LAST = r - 1

U_TOT = U_LAST + 1
put(pu, f"A{U_TOT}", "Total metered usage", BOLD, border=True)
put(pu, f"I{U_TOT}", f"=SUM(I{U_FIRST}:I{U_LAST})", BOLD, fmt=MONEY, border=True)
put(pu, f"J{U_TOT}", f"=SUM(J{U_FIRST}:J{U_LAST})", BOLD, fmt=MONEY, border=True)
band(pu, U_TOT, 1, 11, TOTAL_FILL)
pu.freeze_panes = "A5"
pu.auto_filter.ref = f"A{U_HDR}:K{U_LAST}"

# ================================================================ Summary
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [38, 17, 17, 17, 17, 14, 4]):
    sm.column_dimensions[col].width = w

put(sm, "A1", "Cost summary by product", TITLE)
put(sm, "A2", "Licence costs come from the Licences sheet, usage costs from Platform_Usage.", SUBTITLE)

S_HDR = 4
s_hdrs = ["Product", "Licences / mo", "Metered usage / mo", "Total / mo", "Total / year", "% of total"]
for i, h in enumerate(s_hdrs, start=1):
    put(sm, f"{get_column_letter(i)}{S_HDR}", h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)

S_FIRST = S_HDR + 1
r = S_FIRST
for key, label in SUMMARY_PRODUCTS:
    put(sm, f"A{r}", label, BODY, border=True)
    put(sm, f"B{r}", f'=SUMIF(Licences!$A${L_FIRST}:$A${L_LAST},"{key}",Licences!$G${L_FIRST}:$G${L_LAST})',
        LINK, fmt=MONEY, border=True)
    put(sm, f"C{r}", f'=SUMIF(Platform_Usage!$B${U_FIRST}:$B${U_LAST},"{key}",Platform_Usage!$I${U_FIRST}:$I${U_LAST})',
        LINK, fmt=MONEY, border=True)
    put(sm, f"D{r}", f"=B{r}+C{r}", BODY, fmt=MONEY, border=True)
    put(sm, f"E{r}", f"=D{r}*12", BODY, fmt=MONEY, border=True)
    r += 1
S_LAST = r - 1

S_TOT = S_LAST + 1
put(sm, f"A{S_TOT}", "Total before contingency", BOLD, border=True)
for col in "BCDE":
    put(sm, f"{col}{S_TOT}", f"=SUM({col}{S_FIRST}:{col}{S_LAST})", BOLD, fmt=MONEY, border=True)
band(sm, S_TOT, 1, 6, TOTAL_FILL)
for rr in range(S_FIRST, S_LAST + 1):
    put(sm, f"F{rr}", f"=IFERROR(D{rr}/$D${S_TOT},0)", BODY, fmt=PCT, border=True)
put(sm, f"F{S_TOT}", f"=IFERROR(SUM(F{S_FIRST}:F{S_LAST}),0)", BOLD, fmt=PCT, border=True)

S_CONT = S_TOT + 1
put(sm, f"A{S_CONT}", "Usage contingency (Dashboard input)", BODY, border=True)
put(sm, f"C{S_CONT}", f"=C{S_TOT}*Dashboard!$C$12", BODY, fmt=MONEY, border=True)
put(sm, f"D{S_CONT}", f"=C{S_CONT}", BODY, fmt=MONEY, border=True)
put(sm, f"E{S_CONT}", f"=D{S_CONT}*12", BODY, fmt=MONEY, border=True)

S_GRAND = S_CONT + 1
put(sm, f"A{S_GRAND}", "TOTAL ESTIMATED COST", F(11, True), border=True)
put(sm, f"B{S_GRAND}", f"=B{S_TOT}", F(11, True), fmt=MONEY, border=True)
put(sm, f"C{S_GRAND}", f"=C{S_TOT}+C{S_CONT}", F(11, True), fmt=MONEY, border=True)
put(sm, f"D{S_GRAND}", f"=D{S_TOT}+D{S_CONT}", F(11, True), fmt=MONEY, border=True)
put(sm, f"E{S_GRAND}", f"=D{S_GRAND}*12", F(11, True), fmt=MONEY, border=True)
band(sm, S_GRAND, 1, 6, TOTAL_FILL)

put(sm, f"A{S_GRAND+2}", "Contract term total", BOLD)
put(sm, f"D{S_GRAND+2}", f"=D{S_GRAND}*Dashboard!$C$10", BOLD, fmt=MONEY0, border=True)
put(sm, f"E{S_GRAND+2}", "Monthly total x term months", NOTE)

# ================================================================ Dashboard
db = wb.create_sheet("Dashboard", 1)
db.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKL", [3, 38, 20, 20, 20, 4, 15, 15, 15, 15, 15, 15]):
    db.column_dimensions[col].width = w

put(db, "B2", "GitHub & GitHub Copilot - Cost Estimate Tool", TITLE)
put(db, "B3", "Yellow cells are inputs. Everything else calculates. USD, list prices as at 27 August 2026.", SUBTITLE)

put(db, "B5", "1. Estimate details", SECTION, fill=SECTION_FILL)
band(db, 5, 2, 5, SECTION_FILL, SECTION)
det = [
    ("Customer / organisation", "Contoso Engineering", "text"),
    ("Prepared by", "Your name", "text"),
    ("Estimate date", "2026-08-27", "text"),
    ("GitHub platform plan", "GitHub Enterprise Cloud", "plan"),
    ("Contract term (months)", 12, "int"),
    ("Discount off list (%)", 0.0, "pct"),
    ("Usage contingency (%)", 0.10, "pct"),
]
r = 6
for label, val, kind in det:
    put(db, f"B{r}", label, BODY, border=True)
    fmt = PCT if kind == "pct" else (QTY if kind == "int" else None)
    put(db, f"C{r}", val, INPUT_F, fmt=fmt, fill=INPUT_FILL, border=True)
    r += 1
put(db, "B13", "Billable developers (platform seats)", BODY, border=True)
put(db, "C13", f"=Licences!$E${L_FIRST}", LINK, fmt=QTY, border=True)
put(db, "D6", "Rows 6-8 are labels only - they do not affect the maths.", NOTE)
put(db, "D9", "Drives seat rate and every included allowance.", NOTE)
put(db, "D12", "Buffer applied to metered usage only.", NOTE)

put(db, "B15", "2. Estimate summary", SECTION, fill=SECTION_FILL)
band(db, 15, 2, 5, SECTION_FILL, SECTION)
put(db, "C16", "Per month", HDR, fill=HDR_FILL, align="center", border=True)
put(db, "D16", "Per year", HDR, fill=HDR_FILL, align="center", border=True)
put(db, "E16", "Full term", HDR, fill=HDR_FILL, align="center", border=True)

summ = [
    ("Licences & subscriptions", f"=Summary!$B${S_TOT}"),
    ("Metered platform usage", f"=Summary!$C${S_TOT}"),
    ("Usage contingency", f"=Summary!$C${S_CONT}"),
]
r = 17
for label, formula in summ:
    put(db, f"B{r}", label, BODY, border=True)
    put(db, f"C{r}", formula, LINK, fmt=MONEY0, border=True)
    put(db, f"D{r}", f"=C{r}*12", BODY, fmt=MONEY0, border=True)
    put(db, f"E{r}", f"=C{r}*$C$10", BODY, fmt=MONEY0, border=True)
    r += 1
put(db, f"B{r}", "TOTAL ESTIMATED COST", F(12, True), border=True)
put(db, f"C{r}", f"=SUM(C17:C19)", F(12, True), fmt=MONEY0, border=True)
put(db, f"D{r}", f"=C{r}*12", F(12, True), fmt=MONEY0, border=True)
put(db, f"E{r}", f"=C{r}*$C$10", F(12, True), fmt=MONEY0, border=True)
band(db, r, 2, 5, TOTAL_FILL)
DB_TOTAL_ROW = r

r += 1
put(db, f"B{r}", "Cost per developer per month", BOLD, border=True)
put(db, f"C{r}", f"=IFERROR(C{DB_TOTAL_ROW}/$C$13,0)", BOLD, fmt=MONEY, border=True)
put(db, f"D{r}", "Total cost divided by platform seats.", NOTE)

put(db, "B23", "3. Cost by product", SECTION, fill=SECTION_FILL)
band(db, 23, 2, 5, SECTION_FILL, SECTION)
put(db, "B24", "Product", HDR, fill=HDR_FILL, border=True)
put(db, "C24", "Per month", HDR, fill=HDR_FILL, align="center", border=True)
put(db, "D24", "Per year", HDR, fill=HDR_FILL, align="center", border=True)
put(db, "E24", "% of total", HDR, fill=HDR_FILL, align="center", border=True)

r = 25
for i, (key, label) in enumerate(SUMMARY_PRODUCTS):
    src = S_FIRST + i
    put(db, f"B{r}", label, BODY, border=True)
    put(db, f"C{r}", f"=Summary!$D${src}", LINK, fmt=MONEY, border=True)
    put(db, f"D{r}", f"=C{r}*12", BODY, fmt=MONEY, border=True)
    put(db, f"E{r}", f"=Summary!$F${src}", LINK, fmt=PCT, border=True)
    r += 1
put(db, f"B{r}", "Total before contingency", BOLD, border=True)
put(db, f"C{r}", f"=SUM(C25:C{r-1})", BOLD, fmt=MONEY, border=True)
put(db, f"D{r}", f"=C{r}*12", BOLD, fmt=MONEY, border=True)
put(db, f"E{r}", f"=SUM(E25:E{r-1})", BOLD, fmt=PCT, border=True)
band(db, r, 2, 5, TOTAL_FILL)
DB_PROD_FIRST, DB_PROD_LAST = 25, r - 1

chart = BarChart()
chart.type = "bar"
chart.style = 2
chart.title = "Estimated monthly cost by product"
chart.y_axis.title = None
data = Reference(db, min_col=3, min_row=24, max_row=DB_PROD_LAST)
cats = Reference(db, min_col=2, min_row=DB_PROD_FIRST, max_row=DB_PROD_LAST)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 10.5
chart.width = 19
chart.legend = None
db.add_chart(chart, "G5")

dv = DataValidation(type="list",
                    formula1='"GitHub Free,GitHub Pro,GitHub Free for organizations,GitHub Team,GitHub Enterprise Cloud"',
                    allow_blank=False, showDropDown=False)
db.add_data_validation(dv)
dv.add(db["C9"])

put(db, f"B{r+2}", "Where to type", BOLD)
tips = [
    "Licences sheet - seat counts and active-committer counts.",
    "Platform_Usage sheet - expected monthly usage per SKU.",
    "SKU_Rate_Card sheet - overwrite a list rate to model negotiated pricing.",
    "Plan_Allowances sheet - adjust included entitlements if your agreement differs.",
]
rr = r + 3
for t in tips:
    put(db, f"B{rr}", t, NOTE)
    rr += 1

order = ["Read_Me", "Dashboard", "Licences", "Platform_Usage", "Summary", "SKU_Rate_Card", "Plan_Allowances"]
wb._sheets = [wb[n] for n in order]
wb.active = 1

OUT = sys.argv[1] if len(sys.argv) > 1 else "GitHub_Copilot_Cost_Estimate_Tool.xlsx"
wb.save(OUT)
print("saved", OUT)
print(f"  SKU_Rate_Card    rows {RC_FIRST}-{RC_LAST} ({RC_LAST - RC_FIRST + 1} SKUs)")
print(f"  Licences         rows {L_FIRST}-{L_LAST}, AI credit pool in G{L_POOL}")
print(f"  Platform_Usage   rows {U_FIRST}-{U_LAST} ({U_LAST - U_FIRST + 1} metered SKUs)")
print(f"  Summary          rows {S_FIRST}-{S_LAST}, grand total row {S_GRAND}")
print(f"  Plan_Allowances  selected-plan row {PA_SEL}")